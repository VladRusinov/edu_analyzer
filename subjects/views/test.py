from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from django.http import FileResponse
import io
from openpyxl import load_workbook
from django.db import transaction
import re

from subjects.models import Test, StudentTest, TaskResult
from subjects.serializers import TestSerializer
from subjects.services.analytics import (
    get_test_stats,
    calculate_task_metrics,
    empty_task_statistics,
    task_statistics_response,
    generate_recommendations,
)

class TestViewSet(viewsets.ModelViewSet):
    queryset = Test.objects.all()
    serializer_class = TestSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        subject_id = self.request.query_params.get("subject")
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        return queryset

    @action(detail=True, methods=["get"], url_path="statistics")
    def statistics(self, request, pk=None):
        test = self.get_object()
        mean_score_all_test = get_test_stats(
            test
        )["mean_score"]
        statistics = []

        for task in test.tasks.all():
            metrics = calculate_task_metrics(
                task,
                mean_score_all_test,
            )

            if metrics is None:
                statistics.append(
                    empty_task_statistics(task)
                )
                continue

            statistics.append(
                task_statistics_response(
                    task,
                    metrics,
                )
            )

        return Response({
            "test_id": test.id,
            "test_title": test.title,
            "statistics": statistics,
        })

    @action(detail=True, methods=["get"], url_path="recommendations")
    def recommendations(self, request, pk=None):
        test = self.get_object()
        mean_score_all_test = get_test_stats(
            test
        )["mean_score"]
        recommendations = []

        for task in test.tasks.all():
            metrics = calculate_task_metrics(
                task,
                mean_score_all_test,
            )

            if metrics is None:
                recommendations.append({
                    "task_id": task.id,
                    "task_topic": task.topic,
                    "hours": task.hours,
                    "recommendations": [
                        "Недостаточно данных для формирования рекомендаций"
                    ],
                })
                continue

            recommendations.append({
                "task_id": task.id,
                "task_topic": task.topic,
                "hours": task.hours,
                "mean_score": round(metrics["mean_score"], 4),
                "variance": round(metrics["variance"], 4),
                "uncertainty": round(metrics["uncertainty"], 4),
                "relative_difficulty": round(
                    metrics["relative_difficulty"],
                    4,
                ),
                "recommendations": generate_recommendations(
                    test,
                    metrics,
                ),
            })

        return Response({
            "test_id": test.id,
            "test_title": test.title,
            "recommendations": recommendations,
        })

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        test = self.get_object()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ведомость оценок"

        tasks = test.tasks.all().order_by('id')
        student_tests = test.student_tests.all()

        headers = ["Студент"]
        for task in tasks:
            headers.append(f"{task.topic} ({task.hours}ч)")
        headers.append("Итоговый балл")
        ws.append(headers)

        for st in student_tests:
            row = [st.student_name]

            for task in tasks:
                result_obj = TaskResult.objects.filter(
                    student_test=st,
                    task=task
                ).first()

                if result_obj:
                    row.append(result_obj.result)
                else:
                    row.append("Нет ответа")

            row.append(st.final_score)
            ws.append(row)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = FileResponse(
            buffer,
            as_attachment=True,
            filename=f"test_{test.id}_results.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return response

    @action(detail=True, methods=["post"], url_path="import-excel")
    def import_excel(self, request, pk=None):
        test = self.get_object()
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"error": "Файл не был загружен"},
                status=status.HTTP_400_BAD_REQUEST
            )

        def normalize_header(text):
            if text is None:
                return ""
            text = str(text).strip().lower()
            text = re.sub(r"№\s+", "№", text)
            text = " ".join(text.split())

            return text

        try:
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active
            tasks = list(test.tasks.all().order_by("id"))
            headers = [cell.value for cell in sheet[1]]
            expected_headers = ["Студент"]

            for task in tasks:
                expected_headers.append(
                    f"{task.topic} ({task.hours}ч)"
                )

            expected_headers.append("Итоговый балл")
            normalized_headers = [
                normalize_header(h)
                for h in headers
            ]
            normalized_expected = [
                normalize_header(h)
                for h in expected_headers
            ]

            if normalized_headers != normalized_expected:
                return Response(
                    {
                        "error":
                        "Структура Excel не соответствует шаблону экспорта.",
                        "expected": expected_headers,
                        "received": headers
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            VALID_RESULTS = {
                "Принято",
                "Частично принято",
                "Не принято",
                "Нет ответа"
            }

            with transaction.atomic():

                TaskResult.objects.filter(
                    student_test__test=test
                ).delete()
                test.student_tests.all().delete()
                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ):

                    if not row:
                        continue

                    if not row[0]:
                        continue

                    student_name = str(row[0]).strip()
                    final_score = row[len(tasks) + 1]
                    student_test = StudentTest.objects.create(
                        test=test,
                        student_name=student_name,
                        final_score=float(final_score or 0)
                    )

                    for i, task in enumerate(tasks):
                        result = row[i + 1]
                        if result is None:
                            result = "Нет ответа"

                        result = str(result).strip()

                        if result not in VALID_RESULTS:
                            raise ValueError(
                                f"Недопустимое значение "
                                f"'{result}' "
                                f"у студента "
                                f"'{student_name}' "
                                f"в задании '{task.topic}'"
                            )

                        TaskResult.objects.create(
                            student_test=student_test,
                            task=task,
                            result=result
                        )

            return Response({
                "message": "Импорт выполнен успешно"
            })

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
